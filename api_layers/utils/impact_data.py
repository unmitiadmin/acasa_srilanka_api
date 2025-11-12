import re
from api_layers.models import TblImpactData
from api_lookups.models import LkpCommodity, LkpImpactColor, LkpImpact, LkpClimateScenario, \
    LkpCountry, LkpState, LkpDistrict, LkpIntensityMetric, LkpChangeMetric
from api_layers.exceptions import LayerDataException
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill



class ImpactData:
    def __init__(self, **kwargs):
        self.db = kwargs.get("db")
        self.default_categories = ["Very Low", "Low", "Medium", "High", "Very High", "NA"]
        self.default_nil_label = "No significant impact"
        self.nil_color = "#969696"
        self.commodity_fields = ["c_vlow", "c_low", "c_med", "c_high", "c_vhigh", "c_nil"]
        self.population_fields = ["pop_vlow", "pop_low", "pop_med", "pop_high", "pop_vhigh", "pop_nil"]
        # request body
        self.commodity_id = kwargs.get("commodity_id")
        self.intensity_metric_id = 2 or kwargs.get("intensity_metric_id") # default to Intensity frequency
        self.visualization_scale_id = 1 or kwargs.get("visualization_scale_id") # default to PIXEL level
        self.climate_scenario_id = kwargs.get("climate_scenario_id")
        self.year = kwargs.get("year")
        self.change_metric_id = kwargs.get("change_metric_id")
        self.country_id = kwargs.get("country_id")
        self.state_id = kwargs.get("state_id")
        self.district_id = kwargs.get("district_id")
        self.layer_id = kwargs.get("layer_id")
        # derivative objects
        self.impact_obj = self.db.query(LkpImpact).filter(LkpImpact.id == self.layer_id).first()
        self.climate_scenario_obj = self.db.query(LkpClimateScenario).get(self.climate_scenario_id)
        self.intensity_metric_obj = self.db.query(LkpIntensityMetric).get(self.intensity_metric_id)
        self.change_metric_obj = self.db.query(LkpChangeMetric).get(self.change_metric_id)
        self.commodity_obj = self.db.query(LkpCommodity).filter(LkpCommodity.id == self.commodity_id).first()
        self.lcase = lambda s: re.sub(r'[^a-zA-Z0-9]', '', s).lower()
        self.ucase = lambda s: re.sub(r'[^a-zA-Z0-9]', '', s).upper()
        self.impact_suffix_obj = (
            self.db.query(LkpImpactColor)
            .filter(LkpImpactColor.suffix == f"{self.impact_obj.optcode}_{self.lcase(self.climate_scenario_obj.scenario)}").first()
        )
        

    def get_legend_categories(self):
        commodity = self.commodity_obj.commodity
        if self.impact_obj.impact == "Productivity":
            if self.climate_scenario_obj.scenario == "Baseline":
                if commodity in ["Wheat", "Rice", "Maize"]:
                    return ["Very Low\n<2T", "Low\n2T-3T", "Medium\n3T-4T", "High\n4T-5T", "Very High\n>5T", "NA"]
                if commodity in ["Chickpea", "Millets", "Mustard",  "Soybean", "Sorghum"]:
                    return ["Very Low\n<0.5T", "Low\n0.5T-1T", "Medium\n1T-1.5T", "High\n1.5T-2T", "Very High\n>2T", "NA"]
                return self.default_categories
            else:
                return ["Medium loss\n<-15%", "Low loss\n-15% to -5%", "Nil\n-5% to 5%", "Low gain\n5% to 15%", "Medium gain\n>15%", "NA"]
        if self.impact_obj.impact == "Resilience":
            return ["Very Low\n(>30%)", "Low\n(20-30%)", "Medium\n(10-20%)", "High\n(5-10%)", "Very High\n(<5%)"]
        return self.default_categories


    def get_legend_header(self):
        if self.impact_obj.impact == "Productivity":
            return (
                "Yield (t/ha)" if self.climate_scenario_obj.scenario == "Baseline"
                else "Percent change in yield"
            )
        if self.impact_obj.impact == "Resilience":
            return "Resilience (CV in %)"
        return self.impact_obj.impact


    def get_text_color(self, hex_color):
        hex_color = hex_color.lstrip('#')
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        brightness = 0.299 * r + 0.587 * g + 0.114 * b
        return 'white' if brightness < 128 else 'black'
    

    def get_legend(self):
        if self.climate_scenario_id != 1 and self.year not in [2050, 2080]:
            raise LayerDataException("Please choose the future year, for non baseline data")
        # map_data comes here
        filters = [
            TblImpactData.commodity_id == self.commodity_id,
            TblImpactData.intensity_metric_id == self.intensity_metric_id,
            TblImpactData.visualization_scale_id == self.visualization_scale_id,
            TblImpactData.climate_scenario_id == self.climate_scenario_id,
            TblImpactData.year == self.year,
            TblImpactData.change_metric_id == self.change_metric_id,
            TblImpactData.country_id == self.country_id,
            TblImpactData.state_id == self.state_id,
            TblImpactData.district_id == self.district_id,
            TblImpactData.impact_optcode_id == self.impact_obj.id,
        ]
        map_data = self.db.query(TblImpactData).filter(*filters).first()
        categories = self.get_legend_categories()
        base_categories = self.default_categories
        ramp = self.impact_suffix_obj.ramp
        result = {
            "layer_type": "impact",
            "layer_id": self.layer_id,
            "row_id": map_data.id if map_data else None,
            "header_text": (
                f"Change in {self.get_legend_header()}"
                if self.climate_scenario_id != 1 and self.change_metric_id == 2
                else self.get_legend_header()
            ),
            "footer_text": None,
        }
        result["population_text"] = "Number of rural farm households"
        result["commodity_text"] = (
            f"{self.commodity_obj.commodity} area, hectare (Ha)" if self.commodity_obj.type.type == "Crops"
            else f"{self.commodity_obj.commodity}" if self.commodity_obj.type.type == "Livestock"
            else None
        )
        result["legend"] = []
        for (i, n) in enumerate(categories):
            commodity_value = getattr(map_data, self.commodity_fields[i]) if map_data else None
            population_value = getattr(map_data, self.population_fields[i]) * 0.16 if map_data else None
            base_category = base_categories[i]
            # if not(base_category == "Nil"):
            result["legend"].append({
                "base_category": base_category,
                "named_category": n,
                "color": ramp[i],
                "text_color": self.get_text_color(ramp[i]),
                "commodity_value": commodity_value, 
                "population_value": population_value,
            })
        return result



    def get_table(self):
        if self.climate_scenario_id != 1 and self.year not in [2050, 2080]:
            raise LayerDataException("Please choose the future year, for non baseline data")
        filters = [
            TblImpactData.commodity_id == self.commodity_id,
            TblImpactData.intensity_metric_id == self.intensity_metric_id,
            TblImpactData.visualization_scale_id == self.visualization_scale_id,
            TblImpactData.climate_scenario_id == self.climate_scenario_id,
            TblImpactData.year == self.year,
            TblImpactData.change_metric_id == self.change_metric_id,
            TblImpactData.impact_optcode_id == self.impact_obj.id,
        ]
        map_data = self.db.query(TblImpactData).filter(*filters)
        if not map_data:
            raise LayerDataException("No data available for the selections")
        if self.country_id is None and self.state_id is None and self.district_id is None:
            # Sri Lanka
            selected_location = "Sri Lanka"
            loc_filter = [TblImpactData.state_id == None]
        if self.country_id is not None and self.state_id is None and self.district_id is None:
            # Country only - by states
            selected_location = f"{self.db.query(LkpCountry).get(self.country_id).country}".replace(" ", "_")
            loc_filter = [TblImpactData.country_id == self.country_id,]
        if self.country_id is not None and self.state_id is not None and self.district_id is None:
            # Country - all states and districts
            state_obj = self.db.query(LkpState).get(self.state_id)
            selected_location = f"{state_obj.country.country}_{state_obj.state}".replace(" ", "_")
            loc_filter = [TblImpactData.country_id == self.country_id, TblImpactData.state_id == self.state_id]
        if self.country_id is not None and self.state_id is not None and self.district_id is None:
            # State - all districts
            dist_obj = self.db.query(LkpDistrict).get(self.district_id)
            selected_location = f"{dist_obj.country.country}_{dist_obj.state.state}_{dist_obj.district}".replace(" ", "_")
            loc_filter = [TblImpactData.country_id == self.country_id, TblImpactData.state_id == self.state_id, TblImpactData.district_id == self.district_id]
        map_data = map_data.filter(*loc_filter)
        rows = [{
            "Commodity": row.commodity.commodity,
            "Model": row.intensity_metric.metric,
            "Analysis Level": row.visualization_scale.scale,
            "Scenario": row.climate_scenario.scenario,
            "Year": row.year or "Baseline",
            "Impact": row.impact_optcode.optcode,
            "Metric": row.change_metric.p,
            "Country": row.country.country if row.country else "South Asia",
            "State": row.state.state if row.state else "Total Country",
            "c_vlow": row.c_vlow,
            "c_low": row.c_low,
            "c_med": row.c_med,
            "c_high": row.c_high,
            "c_vhigh": row.c_vhigh,
            "c_nil": row.c_nil,
            "pop_vlow": int(round(row.pop_vlow)) if row.pop_vlow is not None else 0,
            "pop_low": int(round(row.pop_low)) if row.pop_low is not None else 0,
            "pop_med": int(round(row.pop_med)) if row.pop_med is not None else 0,
            "pop_high": int(round(row.pop_high)) if row.pop_high is not None else 0,
            "pop_vhigh": int(round(row.pop_vhigh)) if row.pop_vhigh is not None else 0,
            "pop_nil": int(round(row.pop_nil)) if row.pop_nil is not None else 0,
        } for row in map_data]
        df = pd.DataFrame(rows)
        # Renaming population value header
        df.rename(columns={
            "pop_vlow": f"Rural population under Very Low category",
            "pop_low": f"Rural population under Low category",
            "pop_med": f"Rural population under Medium category",
            "pop_high": f"Rural population under High category",
            "pop_vhigh": f"Rural population under Very High category",
            "pop_nil": f"Rural population under Nil category",
        }, inplace=True)
        commodity_name = self.commodity_obj.commodity
        # Renaming commodity value header
        if self.commodity_obj.type.type == "Crops":
            df.rename(columns={
                "c_vlow": f"Cropped area of {commodity_name} under Very Low category",
                "c_low": f"Cropped area of {commodity_name} under Low category",
                "c_med": f"Cropped area of {commodity_name} under Medium category",
                "c_high": f"Cropped area of {commodity_name} under High category",
                "c_vhigh": f"Cropped area of {commodity_name} under Very High category",
                "c_nil": f"Cropped area of {commodity_name} under Nil category",
            }, inplace=True)
        elif self.commodity_obj.type.type == "Livestock":
            df.rename(columns={
                "c_vlow": f"Number of {commodity_name} under Very Low category",
                "c_low": f"Number of {commodity_name} under Low category",
                "c_med": f"Number of {commodity_name} under Medium category",
                "c_high": f"Number of {commodity_name} under High category",
                "c_vhigh": f"Number of {commodity_name} under Very High category",
                "c_nil": f"Number of {commodity_name} under Nil category",
            }, inplace=True)
        # Save into in-memory file object
        # Maize_DAYKUNDI_CV_Intensity frequency_abs_Baseline_Baseline
        filename = (
            f"{self.commodity_obj.commodity}_"
            f"{selected_location}_"
            f"{self.impact_obj.optcode}_"
            f"{self.intensity_metric_obj.metric}_"
            f"{self.change_metric_obj.metric}_"
            f"{f'{self.year}_{self.ucase(map_data[0].climate_scenario.scenario)}' if self.year else 'Baseline'}"
        ).replace(" ", "_")
        memfile = BytesIO()
        sheet_name = f"{self.commodity_obj.commodity}_{self.impact_obj.optcode}".replace(" ", "_")
        with pd.ExcelWriter(memfile, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            base_categories = self.default_categories
            ramp = self.impact_suffix_obj.ramp
            color_map = {
                f"under {cat} category": PatternFill(start_color=color.replace("#", ""), end_color=color.replace("#", ""), fill_type="solid")
                for cat, color in zip(base_categories, ramp)
            }
            default_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            for cell in worksheet[1]:
                text = str(cell.value or "")
                cell.fill = next((fill for key, fill in color_map.items() if key in text), default_fill)
            worksheet.freeze_panes = "A2"
        memfile.seek(0)
        return {
            "filename": filename,
            "memfile": memfile
        }
