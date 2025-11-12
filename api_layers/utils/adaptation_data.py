import re
from api_layers.models import TblAdaptCropData, TblAdaptLivestockData, TblImpactData
from api_lookups.models import LkpCommodity, LkpClimateScenario, \
    LkpImpactColor, LkpAdapt, LkpAdaptCropColor, LkpAdaptLivestockColor, \
    LkpAdaptCropOptcode, LkpCountry, LkpState, LkpDistrict, LkpIntensityMetric, LkpChangeMetric
from api_layers.exceptions import LayerDataException
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill
from pprint import pprint



class AdaptationData:
    def __init__(self, **kwargs):
        self.db = kwargs.get("db")
        self.default_categories = ["Very Low", "Low", "Medium", "High", "Very High"]
        self.nil_color = "#969696"
        self.commodity_fields = ["c_vlow", "c_low", "c_med", "c_high", "c_vhigh", "c_uns", "c_nil"]
        self.population_fields = ["pop_vlow", "pop_low", "pop_med", "pop_high", "pop_vhigh", "pop_uns", "pop_nil"]
        # request body
        self.adaptation_croptab_id = kwargs.get("adaptation_croptab_id")
        self.commodity_id = kwargs.get("commodity_id")
        self.intensity_metric_id = 2 or kwargs.get("intensity_metric_id") # default to Intensity frequency
        self.visualization_scale_id = 1 or kwargs.get("visualization_scale_id") # default to PIXEL level
        self.climate_scenario_id = kwargs.get("climate_scenario_id")
        self.year = kwargs.get("year") if self.climate_scenario_id != 1 else None
        self.change_metric_id = kwargs.get("change_metric_id")
        self.country_id = kwargs.get("country_id")
        self.state_id = kwargs.get("state_id")
        self.district_id = kwargs.get("district_id")
        self.layer_id = kwargs.get("layer_id")
        # derivative objects
        self.climate_scenario_obj = self.db.query(LkpClimateScenario).get(self.climate_scenario_id)
        self.commodity_obj = self.db.query(LkpCommodity).get(self.commodity_id)
        self.adap_obj = self.db.query(LkpAdapt).get(self.layer_id)
        self.intensity_metric_obj = self.db.query(LkpIntensityMetric).get(self.intensity_metric_id)
        self.change_metric_obj = self.db.query(LkpChangeMetric).get(self.change_metric_id)
        self.lcase = lambda s: re.sub(r'[^a-zA-Z0-9]', '', s).lower()
        self.ucase = lambda s: re.sub(r'[^a-zA-Z0-9]', '', s).upper()



    def get_impact_baseline_productivity_categories(self):
        commodity = self.commodity_obj.commodity
        if commodity in ["Wheat", "Rice", "Maize"]:
            return ["Very Low\n<2T", "Low\n2T-3T", "Medium\n3T-4T", "High\n4T-5T", "Very High\n>5T", "NA"]
        if commodity in ["Chickpea", "Millets", "Mustard",  "Soybean", "Sorghum"]:
            return ["Very Low\n<0.5T", "Low\n0.5T-1T", "Medium\n1T-1.5T", "High\n1.5T-2T", "Very High\n>2T", "NA"]
        return [*self.default_categories, "NA"]


    def get_legend_categories(self):
        scenario = self.climate_scenario_obj.scenario
        commodity = self.commodity_obj
        if commodity.type.type == "Crops":
            crop_cat_map = {
                "Land-climate suitability": ["Unsuitable\narea", "Low", "Medium", "High"],
                "Scalability": [*self.default_categories, "Unsuitable\narea"],
                "Gender suitability": ["Considerable\ndecrease", "Moderate\ndecrease", "No\nchange", "Moderate\nincrease", "Considerable\nincrease", "Unsuitable\narea"],
                "Female labourer suitability": ["Considerable\ndecrease", "Moderate\ndecrease", "No\nchange", "Moderate\nincrease", "Considerable\nincrease", "Unsuitable\narea"],
                "Female cultivator suitability": ["Considerable\ndecrease", "Moderate\ndecrease", "No\nchange", "Moderate\nincrease", "Considerable\nincrease", "Unsuitable\narea"],
                "Yield Benefits": ["Medium loss\n<-15%", "Low loss\n-15% to -5%", "Nil\n-5% to 5%", "Low gain\n5% to 15%", "Medium gain\n>15%", "Unsuitable\narea", "NA"],
                "Economic Viability": [*self.default_categories, "Unsuitable\narea", "NA"],
            }
            adap_croptab_obj = self.db.query(LkpAdaptCropColor).get(self.adaptation_croptab_id)
            if adap_croptab_obj.tab_name == "Adaptation Benefits":
                return (
                    self.get_impact_baseline_productivity_categories() if scenario == "Baseline"
                    else ["Adaptation", "Revisit", "Maladaptation", "Ineffective", "Very High", "Unsuitable\narea", "NA"] # Very High to be excluded
                )
            else:
                return crop_cat_map[adap_croptab_obj.tab_name]
        elif commodity.type.type == "Livestock":
            return ["Inapt", "Very Low", "Low", "Medium", "High"]

    

    def get_legend_bgcolors(self):
        scenario = self.climate_scenario_obj.scenario
        commodity = self.commodity_obj
        if commodity.type.type == "Crops":
            adap_croptab_obj = self.db.query(LkpAdaptCropColor).get(self.adaptation_croptab_id)
            crop_ramp = adap_croptab_obj.ramp
            if adap_croptab_obj.tab_name in ["Yield Benefits", "Economic Viability", "Adaptation Benefits"]:
                # Unsuitable area followed by NA
                crop_ramp[-1], crop_ramp[-2] = crop_ramp[-2], crop_ramp[-1]
            if adap_croptab_obj.tab_name == "Adaptation Benefits":
                crop_ramp.insert(4, "#FFFFFF") # very high placeholder
                return (
                    self.db.query(LkpImpactColor).filter(LkpImpactColor.suffix == "Productivity_baseline").first().ramp 
                    if scenario == "Baseline" else crop_ramp
                )
            else:
                return crop_ramp
        elif commodity.type.type == "Livestock":
            livestock_color_obj = self.db.query(LkpAdaptLivestockColor).filter(LkpAdaptLivestockColor.suffix == self.adap_obj.optcode).first()
            return livestock_color_obj.ramp
    

    def get_legend_header(self):
        scenario = self.climate_scenario_obj.scenario
        commodity = self.commodity_obj
        if commodity.type.type == "Crops":
            adap_croptab_obj = self.db.query(LkpAdaptCropColor).get(self.adaptation_croptab_id)
            return (
                (f"Effectiveness of {self.adap_obj.adaptation}" if scenario != "Baseline" else "Yield (t/ha)") 
                if adap_croptab_obj.tab_name == "Adaptation Benefits"
                else f"Yield benefits due to {self.adap_obj.adaptation}" if adap_croptab_obj.tab_name == "Yield Benefits"
                else f"{adap_croptab_obj.tab_name} for {self.adap_obj.adaptation}"
            )
        elif commodity.type.type == "Livestock":
            return f"Suitability for {self.adap_obj.adaptation}"
        

    def get_text_color(self, hex_color):
        hex_color = hex_color.lstrip('#')
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        brightness = 0.299 * r + 0.587 * g + 0.114 * b
        return 'white' if brightness < 128 else 'black'
    

    def get_legend_values(self):
        if self.commodity_obj.type.type == "Crops":
            # get adaptation_croptab_id from LkpAdaptCropColor (self.adaptation_croptab_id)
            # get adaptation_optcode_id from LkpAdaptCropOptcode (csufx_obj)
            adap_croptab_obj = self.db.query(LkpAdaptCropColor).get(self.adaptation_croptab_id)
            scenario = self.climate_scenario_obj.scenario
            if adap_croptab_obj.tab_name == "Adaptation Benefits" and scenario == "Baseline":
                filters = [
                    TblImpactData.commodity_id == self.commodity_id,
                    TblImpactData.intensity_metric_id == self.intensity_metric_id,
                    TblImpactData.visualization_scale_id == self.visualization_scale_id,
                    TblImpactData.climate_scenario_id == self.climate_scenario_id,
                    TblImpactData.year == self.year,
                    TblImpactData.change_metric_id == self.change_metric_id,
                    TblImpactData.country_id == self.country_id,
                    TblImpactData.state_id == self.state_id,
                    TblImpactData.district_id == self.district_id,
                    TblImpactData.impact_optcode_id == 1, # Productivity X Baseline - one-off exemption
                ]
                map_data = self.db.query(TblImpactData).filter(*filters).first()
                layer_type, layer_id = "impact", 1
            else:
                csufx_obj = self.db.query(LkpAdaptCropOptcode).filter(LkpAdaptCropOptcode.optcode == self.adap_obj.optcode).first()
                filters = [
                    TblAdaptCropData.commodity_id == self.commodity_id,
                    TblAdaptCropData.intensity_metric_id == self.intensity_metric_id,
                    TblAdaptCropData.visualization_scale_id == self.visualization_scale_id,
                    TblAdaptCropData.climate_scenario_id == self.climate_scenario_id,
                    TblAdaptCropData.year == self.year,
                    TblAdaptCropData.change_metric_id == self.change_metric_id,
                    TblAdaptCropData.country_id == self.country_id,
                    TblAdaptCropData.state_id == self.state_id,
                    TblAdaptCropData.district_id == self.district_id,
                    TblAdaptCropData.adaptation_prefix_id == self.adaptation_croptab_id,
                    TblAdaptCropData.adaptation_optcode_id == csufx_obj.id
                ]
                map_data = self.db.query(TblAdaptCropData).filter(*filters).first()
                layer_type, layer_id = "adaptation", self.layer_id
        elif self.commodity_obj.type.type == "Livestock":
            # get adaptation_optcode_id from LkpAdaptLivestockColor (lsufx_obj)
            lsufx_obj = self.db.query(LkpAdaptLivestockColor).filter(LkpAdaptLivestockColor.suffix == self.adap_obj.optcode).first()
            filters = [
                TblAdaptLivestockData.commodity_id == self.commodity_id,
                TblAdaptLivestockData.intensity_metric_id == self.intensity_metric_id,
                TblAdaptLivestockData.visualization_scale_id == self.visualization_scale_id,
                TblAdaptLivestockData.climate_scenario_id == self.climate_scenario_id,
                TblAdaptLivestockData.year == self.year,
                TblAdaptLivestockData.change_metric_id == self.change_metric_id,
                TblAdaptLivestockData.country_id == self.country_id,
                TblAdaptLivestockData.state_id == self.state_id,
                TblAdaptLivestockData.district_id == self.district_id,
                TblAdaptLivestockData.adaptation_optcode_id == lsufx_obj.id
            ]
            map_data = self.db.query(TblAdaptLivestockData).filter(*filters).first()
            layer_type, layer_id = "adaptation", self.layer_id
        return map_data, layer_type, layer_id

    
    def get_legend(self):
        if self.commodity_obj.type.type == "Crops" and not self.adaptation_croptab_id:
            raise LayerDataException("Please choose an appropriate adaptation indicator for the chosen crop")
        if self.climate_scenario_id != 1 and self.year not in [2050, 2080]:
            raise LayerDataException("Please choose the future year, for non baseline data")
        map_data, layer_type, layer_id = self.get_legend_values()
        categories = self.get_legend_categories()
        ramp = [*self.get_legend_bgcolors()]
        result = {
            "layer_type": layer_type,
            "layer_id": layer_id,
            "row_id":  getattr(map_data, "id") if map_data else None,
            "header_text": (
                f"Change in {self.get_legend_header()}"
                if self.climate_scenario_id != 1 and self.change_metric_id == 2
                else self.get_legend_header()
            ),
            "header_bold_part": self.adap_obj.adaptation,
            "footer_text": None,
        }
        result["population_text"] = "Number of rural farm households"
        result["commodity_text"] = (
            f"{self.commodity_obj.commodity} area, hectare (Ha)" if self.commodity_obj.type.type == "Crops"
            else f"{self.commodity_obj.commodity}" if self.commodity_obj.type.type == "Livestock"
            else None
        )
        result["legend"] = []
        scenario = self.climate_scenario_obj.scenario
        adap_croptab_obj = self.db.query(LkpAdaptCropColor).get(self.adaptation_croptab_id)
        for (i, n) in enumerate(categories):
            base_category = categories[i]
            legend_item = {
                "base_category": base_category,
                "named_category": n,
                "color": ramp[i],
                "text_color": self.get_text_color(ramp[i]),
            }
            if self.commodity_obj.type.type == "Crops":
                crop_index = {
                    "Land-climate suitability": [0, 1, 2, 3],
                    "Scalability": [0, 1, 2, 3, 4, 5],
                    "Gender suitability": [0, 1, 2, 3, 4, 5],
                    "Female labourer suitability": [0, 1, 2, 3, 4, 5],
                    "Female cultivator suitability": [0, 1, 2, 3, 4, 5],
                    "Yield Benefits": [0, 1, 2, 3, 4, 6, 5],
                    "Economic Viability": [0, 1, 2, 3, 4, 6, 5],
                    "Adaptation Benefits": [0, 1, 2, 3, 4] if scenario == "Baseline" else [0, 1, 2, 3, 5, 6] # Skipping Very High
                }
                val_index = crop_index[adap_croptab_obj.tab_name]
            else:
                val_index = [0, 1, 2, 3, 4]
            if i in val_index:
                commodity_value = getattr(map_data, self.commodity_fields[i]) if map_data else None
                population_value = getattr(map_data, self.population_fields[i]) * 0.16 if map_data else None
                legend_item.update({
                    "commodity_value": commodity_value,
                    "population_value": population_value,
                })
                result["legend"].append(legend_item)
        if self.commodity_obj.type.type == "Crops" and adap_croptab_obj.tab_name == "Adaptation Benefits" and scenario == "Baseline": # [6 - nil]
            result["legend"].append({
                "base_category": "NA",
                "named_category": "NA",
                "color": self.nil_color,
                "text_color": self.get_text_color(self.nil_color),
                "commodity_value": getattr(map_data, "c_nil") if map_data else None,
                "population_value": getattr(map_data, "pop_nil") if map_data else None,
            })
        return result


    def get_table_values(self):
        TblAdaptData = (
            TblAdaptCropData if self.commodity_obj.type.type == "Crops"
            else TblAdaptLivestockData # Livestock
        )
        if self.commodity_obj.type.type == "Crops":
            adap_croptab_obj = self.db.query(LkpAdaptCropColor).get(self.adaptation_croptab_id)
            scenario = self.climate_scenario_obj.scenario
            if adap_croptab_obj.tab_name == "Adaptation Benefits" and scenario == "Baseline":
                filters = [
                    TblImpactData.commodity_id == self.commodity_id,
                    TblImpactData.intensity_metric_id == self.intensity_metric_id,
                    TblImpactData.visualization_scale_id == self.visualization_scale_id,
                    TblImpactData.climate_scenario_id == self.climate_scenario_id,
                    TblImpactData.year == self.year,
                    TblImpactData.change_metric_id == self.change_metric_id,
                    TblImpactData.impact_optcode_id == 1, # Productivity X Baseline - one-off exemption
                ]
                map_data = self.db.query(TblImpactData).filter(*filters)
                TblAdaptData = TblImpactData
            else:
                csufx_obj = self.db.query(LkpAdaptCropOptcode).filter(LkpAdaptCropOptcode.optcode == self.adap_obj.optcode).first()
                filters = [
                    TblAdaptCropData.commodity_id == self.commodity_id,
                    TblAdaptCropData.intensity_metric_id == self.intensity_metric_id,
                    TblAdaptCropData.visualization_scale_id == self.visualization_scale_id,
                    TblAdaptCropData.climate_scenario_id == self.climate_scenario_id,
                    TblAdaptCropData.year == self.year,
                    TblAdaptCropData.change_metric_id == self.change_metric_id,
                    TblAdaptCropData.adaptation_prefix_id == self.adaptation_croptab_id,
                    TblAdaptCropData.adaptation_optcode_id == csufx_obj.id
                ]
                map_data = self.db.query(TblAdaptCropData).filter(*filters)
                TblAdaptData = TblAdaptCropData
        else:
            lsufx_obj = self.db.query(LkpAdaptLivestockColor).filter(LkpAdaptLivestockColor.suffix == self.adap_obj.optcode).first()
            filters = [
                TblAdaptLivestockData.commodity_id == self.commodity_id,
                TblAdaptLivestockData.intensity_metric_id == self.intensity_metric_id,
                TblAdaptLivestockData.visualization_scale_id == self.visualization_scale_id,
                TblAdaptLivestockData.climate_scenario_id == self.climate_scenario_id,
                TblAdaptLivestockData.year == self.year,
                TblAdaptLivestockData.change_metric_id == self.change_metric_id,
                TblAdaptLivestockData.adaptation_optcode_id == lsufx_obj.id
            ]
            map_data = self.db.query(TblAdaptLivestockData).filter(*filters)
            TblAdaptData = TblAdaptLivestockData
        if not map_data:
            raise LayerDataException("No data available for the selections")
        # Location fields
        if self.country_id is None and self.state_id is None and self.district_id is None:
            # Sri Lanka
            selected_location = "Sri Lanka"
            loc_filter = [TblAdaptData.state_id == None]
        if self.country_id is not None and self.state_id is None and self.district_id is None:
            # Country only - by states
            selected_location = f"{self.db.query(LkpCountry).get(self.country_id).country}".replace(" ", "_")
            loc_filter = [TblAdaptData.country_id == self.country_id]
        if self.country_id is not None and self.state_id is not None and self.district_id is None:
            state_obj = self.db.query(LkpState).get(self.state_id)
            selected_location = f"{state_obj.country.country}_{state_obj.state}".replace(" ", "_")
            loc_filter = [TblAdaptData.country_id == self.country_id, TblAdaptData.state_id == self.state_id]
        if self.country_id is not None and self.state_id is not None and self.district_id is not None:
            dist_obj = self.db.query(LkpDistrict).get(self.district_id)
            selected_location = f"{dist_obj.country.country}_{dist_obj.state.state}_{dist_obj.district}".replace(" ", "_")
            loc_filter = [TblAdaptData.country_id == self.country_id, TblAdaptData.state_id == self.state_id, TblAdaptData.district_id == self.district_id]
        map_data = map_data.filter(*loc_filter)
        return selected_location, map_data

    

    def get_table(self):
        if self.commodity_obj.type.type == "Crops" and not self.adaptation_croptab_id:
            raise LayerDataException("Please choose an appropriate adaptation indicator for the chosen crop")
        if self.climate_scenario_id != 1 and self.year not in [2050, 2080]:
            raise LayerDataException("Please choose the future year, for non baseline data")
        selected_location, map_data = self.get_table_values()
        rows = []

        adap_croptab_obj = self.db.query(LkpAdaptCropColor).get(self.adaptation_croptab_id)
        scenario = self.climate_scenario_obj.scenario

        for row in map_data:
            base = {
                "Commodity": row.commodity.commodity,
                "Model": row.intensity_metric.metric,
                "Analysis Level": row.visualization_scale.scale,
                "Scenario": row.climate_scenario.scenario,
                "Year": row.year or "Baseline",
            }
            if self.commodity_obj.type.type == "Crops":
                if adap_croptab_obj.tab_name == "Adaptation Benefits" and scenario == "Baseline":
                    base["Impact"] = row.impact_optcode.impact
                else:
                    base["Adaptation group"] = row.adaptation_prefix.prefix
                    base["Adaptation"] = row.adaptation_optcode.optcode
            elif self.commodity_obj.type.type == "Livestock":
                base["Adaptation"] = row.adaptation_optcode.suffix
            base["Metric"] = row.change_metric.metric
            base["Country"] = row.country.country if row.country else "South Asia"
            base["State"] = row.state.state if row.state else ("Total Country" if row.country else "All countries")
            base["c_vlow"] = row.c_vlow
            base["c_low"] = row.c_low
            base["c_med"] = row.c_med
            base["c_high"] = row.c_high
            base["c_vhigh"] = row.c_vhigh
            if (
                self.commodity_obj.type.type == "Crops" 
                and not(adap_croptab_obj.tab_name == "Adaptation Benefits" and scenario == "Baseline")
            ): base["c_uns"] = row.c_uns
            base["c_nil"] = row.c_nil
            base["pop_vlow"] = row.pop_vlow
            base["pop_low"] = row.pop_low
            base["pop_med"] = row.pop_med
            base["pop_high"] = row.pop_high
            base["pop_vhigh"] = row.pop_vhigh
            if (
                self.commodity_obj.type.type == "Crops" 
                and not(adap_croptab_obj.tab_name == "Adaptation Benefits" and scenario == "Baseline")
            ): base["pop_uns"] = row.pop_uns
            base["pop_nil"] = row.pop_nil
            rows.append(base)
        df = pd.DataFrame(rows)
        # Renaming population value header
        df.rename(columns={
            "pop_vlow": f"Rural population under Very Low suitability",
            "pop_low": f"Rural population under Low suitability",
            "pop_med": f"Rural population under Medium suitability",
            "pop_high": f"Rural population under High suitability",
            "pop_vhigh": f"Rural population under Very High suitability",
            "pop_uns": f"Rural population under Unsuitable area",
            "pop_nil": f"Rural population under Nil suitability",
        }, inplace=True)
        commodity_name = self.commodity_obj.commodity
        # Renaming commodity value header
        if self.commodity_obj.type.type == "Crops":
            df.rename(columns={
                "c_vlow": f"Cropped area of {commodity_name} under Very Low suitability",
                "c_low": f"Cropped area of {commodity_name} under Low suitability",
                "c_med": f"Cropped area of {commodity_name} under Medium suitability",
                "c_high": f"Cropped area of {commodity_name} under High suitability",
                "c_vhigh": f"Cropped area of {commodity_name} under Very High suitability",
                "c_uns": f"Cropped area of {commodity_name} under Unsuitable area",
                "c_nil": f"Cropped area of {commodity_name} under Nil suitability",
            }, inplace=True)
        elif self.commodity_obj.type.type == "Livestock":
            df.rename(columns={
                "c_vlow": f"Number of {commodity_name} under Very Low suitability",
                "c_low": f"Number of {commodity_name} under Low suitability",
                "c_med": f"Number of {commodity_name} under Medium suitability",
                "c_high": f"Number of {commodity_name} under High suitability",
                "c_vhigh": f"Number of {commodity_name} under Very High suitability",
                "c_uns": f"Number of {commodity_name} under Unsuitable area",
                "c_nil": f"Number of {commodity_name} under Nil suitability",
            }, inplace=True)
        # Save into in-memory file object
        filename = (
            f"{self.commodity_obj.commodity}_"
            f"{selected_location}_"
            f"{self.adap_obj.adaptation}_"
            f"{self.intensity_metric_obj.metric}_"
            f"{self.change_metric_obj.metric}_"
            f"{f'{self.year}_{self.ucase(self.climate_scenario_obj.scenario)}' if self.year else 'Baseline'}"
        ).replace(" ", "_")
        memfile = BytesIO()
        sheet_name = f"{self.commodity_obj.commodity}_{self.adap_obj.adaptation}".replace(" ", "_")
        with pd.ExcelWriter(memfile, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            ramp = [str(c).lstrip("#")[:6].upper() for c in self.get_legend_bgcolors()]
            fills = [
                PatternFill(start_color=color, end_color=color, fill_type="solid")
                for color in ramp
            ]
            default_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            commodity_cells = []
            population_cells = []
            other_cells = []
            for cell in worksheet[1]:
                text = str(cell.value or "").lower()
                if "under" in text:
                    if "population" in text:
                        population_cells.append(cell)
                    else:
                        commodity_cells.append(cell)
                else:
                    other_cells.append(cell)
            # Assign ramp colors to commodity headers
            for i, cell in enumerate(commodity_cells):
                cell.fill = fills[i] if i < len(fills) else default_fill
            # Assign ramp colors again (restart index) to population headers
            for i, cell in enumerate(population_cells):
                cell.fill = fills[i] if i < len(fills) else default_fill
            # Everything else → default
            for cell in other_cells:
                cell.fill = default_fill
        memfile.seek(0)
        return {
            "filename": filename,
            "memfile": memfile
        }

