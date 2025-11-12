import re
from api_layers.models import TblRiskData
from api_lookups.models import LkpCommodity, LkpRisk, LkpRiskColor, LkpCountry, LkpState, LkpDistrict, \
    LkpChangeMetric, LkpIntensityMetric
from api_layers.exceptions import LayerDataException
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill
from sqlalchemy.dialects import mysql


class RiskData:
    def __init__(self, **kwargs):
        self.db = kwargs.get("db")
        self.default_categories = ["Very Low", "Low", "Medium", "High", "Very High"]
        self.default_nil_label = "No significant risk"
        self.nil_color = "#969696"
        self.commodity_fields = ["c_nil", "c_vlow", "c_low", "c_med", "c_high", "c_vhigh"]
        self.population_fields = ["pop_nil", "pop_vlow", "pop_low", "pop_med", "pop_high", "pop_vhigh"]
        # request body
        self.commodity_id = kwargs.get("commodity_id")
        self.intensity_metric_id = 2 or kwargs.get("intensity_metric_id") # default to Intensity frequency
        self.visualization_scale_id = 1 or kwargs.get("visualization_scale_id") # default to PIXEL level
        self.analysis_scope_id = kwargs.get("analysis_scope_id") # Do not consider commodities when 2 (take regional analysis)
        self.climate_scenario_id = kwargs.get("climate_scenario_id")
        self.year = kwargs.get("year")
        self.change_metric_id = kwargs.get("change_metric_id")
        self.country_id = kwargs.get("country_id")
        self.state_id = kwargs.get("state_id")
        self.district_id = kwargs.get("district_id")
        self.layer_id = kwargs.get("layer_id")
        # derivative objects
        self.risk_obj = self.db.query(LkpRisk).filter(LkpRisk.id == self.layer_id).first()
        self.risk_suffix_obj = self.db.query(LkpRiskColor).filter(LkpRiskColor.suffix == self.risk_obj.suffix).first()
        self.commodity_obj = self.db.query(LkpCommodity).filter(LkpCommodity.id == self.commodity_id).first()
        self.intensity_metric_obj = self.db.query(LkpIntensityMetric).get(self.intensity_metric_id)
        self.change_metric_obj = self.db.query(LkpChangeMetric).get(self.change_metric_id)
        self.lcase = lambda s: re.sub(r'[^a-zA-Z0-9]', '', s).lower()
        self.ucase = lambda s: re.sub(r'[^a-zA-Z0-9]', '', s).upper()


    def get_legend_categories(self):
        suffix = self.risk_suffix_obj.suffix
        if self.analysis_scope_id == 1:
            commodity = self.commodity_obj
            if suffix == "Seasonal Rainfall":
                if self.change_metric_id == 1: # abs
                    if commodity.type.type == "Crops":
                        if commodity.commodity in ["Wheat", "Sorghum", "Chickpea", "Lentil", "Mustard"]:
                            return ["<20mm", "20-50mm", "50-75mm", "75-100mm", ">100mm"]
                        else:
                            return ["<250mm", "250-500mm", "500-750mm", "750-1000mm", ">1000mm"]
                    elif commodity.type.type == "Livestock":
                        return ["<500mm", "500-1000mm", "1000-1500mm", "1500-2000mm", ">2000mm"]
                elif self.change_metric_id == 2: # del
                    return ["<-100mm", "-100-50mm", "-50-50mm", "50-100mm", ">100mm"]
            if suffix == "Maximum Temperature":
                if self.change_metric_id == 1: # abs
                    if commodity.type.type == "Crops":
                        if commodity.commodity in ["Wheat", "Sorghum", "Chickpea", "Lentil", "Mustard"]:
                            return ["<16°C", "16-20°C", "20-24°C", "24-28°C", ">28°C"]
                        else:
                            return ["<27.5°C", "27.5-30°C", "30-32.5°C", "32.5-35°C", ">35°C"]
                    elif commodity.type.type == "Livestock":
                        return ["<25°C", "25-27.5°C", "27.5-30°C", "30-32.5°C", ">32.5°C"]
                elif self.change_metric_id == 2: # del
                    return ["0-0.5°C", "0.5-1°C", "1-1.5°C", "1.5-2°C", ">2°C"]
            if suffix == "Minimum Temperature":
                if self.change_metric_id == 1: # abs    
                    if commodity.type.type == "Crops":
                        if commodity.commodity in ["Wheat", "Sorghum", "Chickpea", "Lentil", "Mustard"]:
                            return ["<10°C", "10-12°C", "12-14°C", "14-16°C", ">16°C"]
                        else:
                            return ["<19°C", "19-21°C", "30-32.5°C", "32.5-35°C", ">35°C"]
                    elif commodity.type.type == "Livestock":
                        return ["<12.5°C", "12.5-15°C", "15-17.5°C", "17.5-20°C", ">20°C"]
                elif self.change_metric_id == 2: # del
                    return ["0-1°C", "1-2°C", "2-3°C", "3-4°C", ">4°C"]
        return (
            # Non climatology layers
            self.default_categories if self.change_metric_id == 1
            else ["Considerable\ndecrease", "Moderate\ndecrease", "No\nchange", "Moderate\nincrease", "Considerable\nincrease"]
        )
    

    def get_text_color(self, hex_color):
        hex_color = hex_color.lstrip('#')
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        brightness = 0.299 * r + 0.587 * g + 0.114 * b
        return 'white' if brightness < 128 else 'black'


    def get_legend(self):
        if self.climate_scenario_id != 1 and self.year not in [2050, 2080]:
            raise LayerDataException("Please choose the future year, for non baseline data")
        filters = [
            TblRiskData.commodity_id == (None if self.analysis_scope_id == 2 else self.commodity_id),
            TblRiskData.intensity_metric_id == self.intensity_metric_id,
            TblRiskData.visualization_scale_id == self.visualization_scale_id,
            TblRiskData.climate_scenario_id == self.climate_scenario_id,
            TblRiskData.year == self.year,
            TblRiskData.change_metric_id == self.change_metric_id,
            TblRiskData.country_id == self.country_id,
            TblRiskData.state_id == self.state_id,
            TblRiskData.district_id == self.district_id,
            TblRiskData.risk_suffix_id == self.risk_suffix_obj.id,
        ]
        map_data = self.db.query(TblRiskData).filter(*filters).first()
        # if not map_data:
        #     return None
        categories = ["Nil", *self.get_legend_categories()]
        base_categories = ["Nil", *self.default_categories]
        delta_color_ramp = (
            ["#800000", "#FF6666", "#F5F5DC", "#87CEFA", "#4682B4"]
            if self.risk_obj.ipcc.ipcc == "Climatology" and self.risk_obj.suffix == "Seasonal Rainfall"
            else ["#4682B4", "#87CEFA", "#E7E6A8", "#FF6666", "#800000"]
        )
        ramp = (
            [self.nil_color, *delta_color_ramp] 
            if self.change_metric_id == 2 and self.climate_scenario_id != 1
            else [self.nil_color, *self.risk_suffix_obj.ramp] 
        )
        result = {
            "layer_type": "risk",
            "layer_id": self.layer_id,
            "row_id": map_data.id if map_data else None,
            "header_text": (
                f"Change in {self.risk_obj.risk}"
                if self.climate_scenario_id != 1 and self.change_metric_id == 2
                else self.risk_obj.risk
            ),
            "footer_text": (
               f"(Lower {str(self.risk_obj.risk).lower()} depicts higher vulnerability)" 
               if self.risk_obj.ipcc.ipcc == "Vulnerability"
               else None
            ),
        }
        result["population_text"] = "Number of rural farm households"
        result["commodity_text"] =( 
            (
                f"{self.commodity_obj.commodity} area, hectare (Ha)" if self.commodity_obj.type.type == "Crops"
                else f"{self.commodity_obj.commodity}" if self.commodity_obj.type.type == "Livestock"
                else None
            ) if self.analysis_scope_id == 1
            else f"Area under regional analysis, hectare (Ha)"
        )
        result["legend"] = []
        for (i, n) in enumerate(categories):
            commodity_value = getattr(map_data, self.commodity_fields[i]) if map_data else None
            population_value = getattr(map_data, self.population_fields[i]) * 0.16 if map_data else None
            base_category = base_categories[i]
            if not(base_category == "Nil" and not (commodity_value or population_value)):
                result["legend"].append({
                    "base_category": base_category,
                    "named_category": n,
                    "color": ramp[i],
                    "text_color": self.get_text_color(ramp[i]),
                    "commodity_value": commodity_value, 
                    "population_value": population_value,
                })
        return result


    def get_table(self):
        if self.climate_scenario_id != 1 and self.year not in [2050, 2080]:
            raise LayerDataException("Please choose the future year, for non baseline data")
        filters = [
            TblRiskData.commodity_id == (None if self.analysis_scope_id == 2 else self.commodity_id),
            TblRiskData.intensity_metric_id == self.intensity_metric_id,
            TblRiskData.visualization_scale_id == self.visualization_scale_id,
            TblRiskData.climate_scenario_id == self.climate_scenario_id,
            TblRiskData.year == self.year,
            TblRiskData.change_metric_id == self.change_metric_id,
            TblRiskData.risk_suffix_id == self.risk_suffix_obj.id,
        ]
        map_data = self.db.query(TblRiskData).filter(*filters)
        if not map_data:
            raise LayerDataException("No data available for the selections")
        if self.country_id is None and self.state_id is None and self.district_id is None:
            # Sri Lanka
            selected_location = "Sri Lanka"
            loc_filter = [TblRiskData.state_id == None]
        if self.country_id is not None and self.state_id is None and self.district_id is None:
            # Country only - by states
            selected_location = f"{self.db.query(LkpCountry).get(self.country_id).country}".replace(" ", "_")
            loc_filter = [TblRiskData.country_id == self.country_id,]
        if self.country_id is not None and self.state_id is not None and self.district_id is None:
            # Country - all states and districts
            state_obj = self.db.query(LkpState).get(self.state_id)
            selected_location = f"{state_obj.country.country}_{state_obj.state}".replace(" ", "_")
            loc_filter = [TblRiskData.country_id == self.country_id, TblRiskData.state_id == self.state_id]
        if self.country_id is not None and self.state_id is not None and self.district_id is not None:
            # Specific district
            dist_obj = self.db.query(LkpDistrict).get(self.district_id)
            selected_location = f"{dist_obj.country.country}_{dist_obj.state.state}_{dist_obj.district}".replace(" ", "_")
            loc_filter = [TblRiskData.country_id == self.country_id, TblRiskData.state_id == self.state_id, TblRiskData.district_id == self.district_id]
        map_data = map_data.filter(*loc_filter)
        rows = [{
            "Commodity": row.commodity.commodity if row.commodity and self.analysis_scope_id == 1 else "Regional",
            "Model": row.intensity_metric.metric,
            "Analysis Level": row.visualization_scale.scale,
            "Scenario": row.climate_scenario.scenario,
            "Year": row.year or "Baseline",
            "Hazard": row.risk_suffix.suffix,
            "Metric": row.change_metric.metric,
            "Country": row.country.country,
            "State": row.state.state if row.state else "Total Country",
            "District": row.district.district if row.district else ("Total State" if row.state else "Total Country"),
            "c_vlow": row.c_vlow,
            "c_low": row.c_low,
            "c_med": row.c_med,
            "c_high": row.c_high,
            "c_vhigh": row.c_vhigh,
            "c_nil": row.c_nil,
            "pop_vlow": int(round(row.pop_vlow)) if row.pop_vlow is not None else 0,
            "pop_low": int(round(row.pop_low)) if row.pop_low is not None else 0,
            "pop_med": int(round(row.pop_med)) if row.pop_med is not None else 0,
            "pop_high": int(round(row.pop_high)) if row.pop_high is not None else 0,
            "pop_vhigh": int(round(row.pop_vhigh)) if row.pop_vhigh is not None else 0,
            "pop_nil": int(round(row.pop_nil)) if row.pop_nil is not None else 0,
        } for row in map_data]
        df = pd.DataFrame(rows)
        # Renaming population value header
        df.rename(columns={
            "pop_vlow": f"Rural population under Very Low category",
            "pop_low": f"Rural population under Low category",
            "pop_med": f"Rural population under Medium category",
            "pop_high": f"Rural population under High category",
            "pop_vhigh": f"Rural population under Very High category",
            "pop_nil": f"Rural population under Nil category",
        }, inplace=True)
        if self.analysis_scope_id == 1:
            commodity_name = self.commodity_obj.commodity
            # Renaming commodity value header
            if self.commodity_obj.type.type == "Crops":
                df.rename(columns={
                    "c_vlow": f"Cropped area of {commodity_name} under Very Low category",
                    "c_low": f"Cropped area of {commodity_name} under Low category",
                    "c_med": f"Cropped area of {commodity_name} under Medium category",
                    "c_high": f"Cropped area of {commodity_name} under High category",
                    "c_vhigh": f"Cropped area of {commodity_name} under Very High category",
                    "c_nil": f"Cropped area of {commodity_name} under Nil category",
                }, inplace=True)
            elif self.commodity_obj.type.type == "Livestock":
                df.rename(columns={
                    "c_vlow": f"Number of {commodity_name} under Very Low category",
                    "c_low": f"Number of {commodity_name} under Low category",
                    "c_med": f"Number of {commodity_name} under Medium category",
                    "c_high": f"Number of {commodity_name} under High category",
                    "c_vhigh": f"Number of {commodity_name} under Very High category",
                    "c_nil": f"Number of {commodity_name} under Nil category",
                }, inplace=True)
        else:
            df.rename(columns={
                "c_vlow": f"Area under Very Low category",
                "c_low": f"Area under Low category",
                "c_med": f"Area under Medium category",
                "c_high": f"Area under High category",
                "c_vhigh": f"Area under Very High category",
                "c_nil": f"Area under Nil category",
            })
        # Save into in-memory file object
        # Barley_BADAKHSHAN_Agricultural GDP_Intensity frequency_abs_Baseline
        filename = (
            f"{self.commodity_obj.commodity if self.analysis_scope_id == 1 else 'Regional'}_"
            f"{selected_location}_"
            f"{self.risk_suffix_obj.suffix}_"
            f"{self.intensity_metric_obj.metric}_"
            f"{self.change_metric_obj.metric}_"
            f"{f'{self.year}_{self.ucase(map_data[0].climate_scenario.scenario)}' if self.year else 'Baseline'}"
        ).replace(" ", "_")
        memfile = BytesIO()
        sheet_name = (
            f"{self.commodity_obj.commodity}_{self.risk_suffix_obj.suffix}".replace(" ", "_")
            if self.analysis_scope_id == 1
            else f"Baseline_{self.risk_suffix_obj.suffix}".replace(" ", "_")
        )
        with pd.ExcelWriter(memfile, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            base_categories = ["Nil", *self.default_categories]
            delta_color_ramp = (
                ["#800000", "#FF6666", "#F5F5DC", "#87CEFA", "#4682B4"]
                if self.risk_obj.ipcc.ipcc == "Climatology" and self.risk_obj.suffix == "Seasonal Rainfall"
                else ["#4682B4", "#87CEFA", "#E7E6A8", "#FF6666", "#800000"]
            )
            ramp = (
                [self.nil_color, *delta_color_ramp] 
                if self.change_metric_id == 2 and self.climate_scenario_id != 1
                else [self.nil_color, *self.risk_suffix_obj.ramp] 
            )
            color_map = {
                f"under {cat} category": PatternFill(start_color=color.replace("#", ""), end_color=color.replace("#", ""), fill_type="solid")
                for cat, color in zip(base_categories, ramp)
            }
            default_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            for cell in worksheet[1]:
                text = str(cell.value or "")
                cell.fill = next((fill for key, fill in color_map.items() if key in text), default_fill)
            worksheet.freeze_panes = "A2"
        memfile.seek(0)
        return {
            "filename": filename,
            "memfile": memfile
        }

        